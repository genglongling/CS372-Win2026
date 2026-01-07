#!/usr/bin/env python3
"""
Script to add/verify columns in Studentlist.xlsx
Adds: First Name, Last Name, Email columns if they don't exist
"""

import sys
import os

# Path to the Excel file (in root directory)
excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Studentlist.xlsx')

try:
    import openpyxl
    
    if not os.path.exists(excel_path):
        print(f"Creating new Excel file: {excel_path}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
    else:
        print(f"Opening existing Excel file: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    
    # Check current headers
    headers = []
    if ws.max_row > 0:
        for cell in ws[1]:
            headers.append(cell.value if cell.value else "")
    
    print(f"\nCurrent headers: {headers}")
    
    # Define required columns
    required_columns = ["First Name", "Last Name", "Email"]
    
    # Check if headers exist
    if not headers or len(headers) == 0:
        # No headers, add them
        print("\nNo headers found. Adding headers...")
        for col_idx, col_name in enumerate(required_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        print(f"Added headers: {required_columns}")
    else:
        # Check which columns are missing
        missing_columns = []
        for col in required_columns:
            if col not in headers:
                missing_columns.append(col)
        
        if missing_columns:
            print(f"\nMissing columns: {missing_columns}")
            # Add missing columns
            next_col = len(headers) + 1
            for col in missing_columns:
                ws.cell(row=1, column=next_col, value=col)
                print(f"Added column '{col}' at column {next_col}")
                next_col += 1
        else:
            print("\nAll required columns already exist!")
    
    # Save the file
    wb.save(excel_path)
    print(f"\n✓ Excel file updated successfully: {excel_path}")
    print("\nCurrent structure:")
    print("  Column 1: First Name")
    print("  Column 2: Last Name")
    print("  Column 3: Email")
    print("\nYou can now add student data to the Excel file.")
    
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it using: pip install openpyxl")
    print("\nOr run: pip3 install openpyxl")
    sys.exit(1)
except Exception as e:
    print(f"Error processing Excel file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

