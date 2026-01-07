#!/usr/bin/env python3
"""
Script to assign students to 20 groups for Assignment 1
Each BenchmarkT3-BucketLarge file is assigned to 2 groups for cross-validation
"""

import csv
import random
import os
import sys

# Path to files (in root directory)
root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
csv_path = os.path.join(root_dir, 'Studentlist.csv')
excel_path = os.path.join(root_dir, 'Studentlist.xlsx')

# Read students from CSV or Excel
students = []

# Try CSV first
if os.path.exists(csv_path):
    try:
        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                first_name = row.get('First Name', '').strip()
                last_name = row.get('Last Name', '').strip()
                email = row.get('Email', '').strip()
                if first_name and last_name:
                    students.append({
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'full_name': f"{first_name} {last_name}"
                    })
        print(f"Read {len(students)} students from CSV file")
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        students = []

# If CSV failed or doesn't exist, try Excel
if not students and os.path.exists(excel_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Find column indices
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col_idx
        
        first_name_col = headers.get('First Name')
        last_name_col = headers.get('Last Name')
        email_col = headers.get('Email')
        
        if first_name_col and last_name_col:
            for row_idx in range(2, ws.max_row + 1):
                first_name = ws.cell(row=row_idx, column=first_name_col).value
                last_name = ws.cell(row=row_idx, column=last_name_col).value
                email = ws.cell(row=row_idx, column=email_col).value if email_col else ""
                
                if first_name and last_name:
                    students.append({
                        'first_name': str(first_name).strip(),
                        'last_name': str(last_name).strip(),
                        'email': str(email).strip() if email else "",
                        'full_name': f"{str(first_name).strip()} {str(last_name).strip()}"
                    })
            print(f"Read {len(students)} students from Excel file")
        else:
            print("Error: Required columns (First Name, Last Name) not found in Excel file")
    except ImportError:
        print("Error: openpyxl not installed. Cannot read Excel file.")
        print("Please install it using: pip3 install openpyxl")
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if not students:
    print(f"\nError: Could not read students from {csv_path} or {excel_path}")
    print("Please ensure one of these files exists with First Name, Last Name, and Email columns")
    sys.exit(1)

print(f"Total students: {len(students)}")
print(f"Total groups: 20")
print(f"Students per group: {len(students) // 20}")
if len(students) % 20 != 0:
    print(f"Note: {len(students) % 20} groups will have 1 extra student")

# Shuffle students for random assignment
random.seed(42)  # Set seed for reproducibility
random.shuffle(students)

# Assign students to groups
# Each BenchmarkT3-BucketLarge file (1-10) gets 2 groups
groups = {}
for i in range(1, 21):
    groups[i] = []

# Distribute students evenly
for idx, student in enumerate(students):
    group_num = (idx % 20) + 1
    groups[group_num].append(student)

# Determine which BenchmarkT3-BucketLarge file each group works on
# Groups 1-2: file 1, Groups 3-4: file 2, etc.
benchmark_assignments = {}
for group_num in range(1, 21):
    benchmark_num = ((group_num - 1) // 2) + 1
    benchmark_assignments[group_num] = benchmark_num

# Print group assignments
print("\n" + "="*80)
print("GROUP ASSIGNMENTS")
print("="*80)

for group_num in sorted(groups.keys()):
    benchmark_file = f"BenchmarkT3-BucketLarge-{benchmark_assignments[group_num]}.tex"
    print(f"\nGroup {group_num}: {benchmark_file} ({len(groups[group_num])} students)")
    print("-" * 80)
    for student in groups[group_num]:
        print(f"  - {student['full_name']} ({student['email']})")

# Generate CSV output
output_csv = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'group_assignments.csv')
with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['Group', 'Benchmark File', 'First Name', 'Last Name', 'Email'])
    for group_num in sorted(groups.keys()):
        benchmark_file = f"BenchmarkT3-BucketLarge-{benchmark_assignments[group_num]}.tex"
        for student in groups[group_num]:
            writer.writerow([
                group_num,
                benchmark_file,
                student['first_name'],
                student['last_name'],
                student['email']
            ])

print(f"\n✓ Group assignments saved to: {output_csv}")

# Print summary by benchmark file
print("\n" + "="*80)
print("SUMMARY BY BENCHMARK FILE")
print("="*80)
for benchmark_num in range(1, 11):
    benchmark_file = f"BenchmarkT3-BucketLarge-{benchmark_num}.tex"
    group1 = benchmark_num * 2 - 1
    group2 = benchmark_num * 2
    total_students = len(groups[group1]) + len(groups[group2])
    print(f"\n{benchmark_file}:")
    print(f"  Group {group1}: {len(groups[group1])} students")
    print(f"  Group {group2}: {len(groups[group2])} students")
    print(f"  Total: {total_students} students")

print("\n" + "="*80)
print("CROSS-VALIDATION PAIRS FOR ASSIGNMENT 2")
print("="*80)
for benchmark_num in range(1, 11):
    group1 = benchmark_num * 2 - 1
    group2 = benchmark_num * 2
    print(f"  Groups {group1} and {group2} will swap datasets in Assignment 2")

